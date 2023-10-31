from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import time

# Ruta al controlador del navegador (ejemplo para Chrome)
driver_path = 'C:\\Users\\pispa\\chromedriver_win32\\chromedriver.exe'  # Cambia esto a la ruta de tu controlador

# Configura las opciones del controlador
chrome_options = webdriver.ChromeOptions()
#chrome_options.add_argument('--headless')  # Habilita el modo sin cabeza

chrome_options.binary_location = r'C:\Program Files\Google\Chrome\Application\chrome.exe'  # Cambia esto a la ruta de tu navegador Chrome

# URL de la página web


# Inicializa el navegador web
chrome_options.add_argument('--incognito')  # Agrega la opción para iniciar en modo incógnito
chrome_options.add_argument('--headless')  # Habilita el modo sin cabeza
driver = webdriver.Chrome(options=chrome_options)
driver.maximize_window()
url = 'https://www.facebook.com'

# Abre la página web
driver.get(url)

#Busca las cookies y las acepta
boton_permitir_cookies = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='cookie-policy-manage-dialog-accept-button']"))
)

boton_permitir_cookies.click()
# Credenciales

usuario = '********************'
contrasena = '********'


# Encuentra los campos de entrada del usuario y la contraseña
input_usuario = WebDriverWait(driver,10).until(EC.presence_of_element_located((By.NAME, 'email')))  # Reemplaza 'usuario' con el nombre real del campo
input_contrasena = driver.find_element(By.NAME, 'pass')  # Reemplaza 'password' con el nombre real del campo

# Ingresa las credenciales en los campos
input_usuario.send_keys(usuario)
input_contrasena.send_keys(contrasena)
input_contrasena.submit()
nombreBuscado='saltando la dieta'
input_buscador=WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//input[@placeholder='Search Facebook']")))
input_buscador.send_keys(nombreBuscado)
input_buscador.send_keys(Keys.ENTER)

# Dejamos cargar la pagina y buscamos la publicación
time.sleep(30)
driver.implicitly_wait(10)
nombreSaltandolaDieta=WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//span[text()='Saltando la Dieta']")))
nombreSaltandolaDieta.click()
foto=WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//img[@class='x1rg5ohu x5yr21d xl1xv1r xh8yej3']")))
driver.execute_script("arguments[0].scrollIntoView(true);", foto)
foto.click()
driver.implicitly_wait(10)

Comentarios =WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Más pertinentes')]")))

# Realiza un desplazamiento hacia el elemento
actions = ActionChains(driver)
actions.move_to_element(Comentarios).perform()
time.sleep(1)  # Espera 1 segundo para dar tiempo al desplazamiento

# Hace clic en el elemento

Comentarios.click()
time.sleep(1)
todoslosComentarios=WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//span[contains(text(),'Todos los comentarios')]")))
todoslosComentarios.click()

error=0
# Con este bucle cargamos todos los comentarios
for i in range(68):  # Itera 68 veces (ajusta según tus necesidades)
    
    try:
        driver.implicitly_wait(2)
        MasComentarios = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//span[contains(text(), 'Ver más comentarios')]")))

        driver.execute_script("arguments[0].scrollIntoView(true);", MasComentarios)
        MasComentarios.click()
        time.sleep(5)
        error=0
    except Exception as e:
        time.sleep(20)
        print('Entro espera 20 seg')
        error=error+1
        if error==2:
            print("No se encontraron más comentarios o se produjo un error:", str(e))
            break  # Sale del bucle si no se pueden cargar más comentarios

time.sleep(10)
driver.implicitly_wait(10)
body = driver.find_element(By.TAG_NAME,'body')
print('Guardando Datos')

# Guardamos los comentarios en un excel

for i in range(6):  # Puedes ajustar el número de desplazamientos necesarios
    body.send_keys(Keys.PAGE_DOWN)
    time.sleep(2)  # Espera a que se carguen los elementos

comment_elements = driver.find_elements(By.XPATH,"//div[@class='x1y1aw1k xn6708d xwib8y2 x1ye3gou']")

# Crear un nuevo archivo de Excel
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Comentarios"

# Añadir encabezados
sheet.cell(row=1, column=1, value="Nombre del Usuario")
sheet.cell(row=1, column=2, value="Texto del Comentario")


# Itera a través de los comentarios y guarda los datos en Excel
row = 2  # Empezar desde la segunda fila
for comment_element in comment_elements:
    user_name = comment_element.find_element(By.XPATH,".//a[@role='link']").text
    comment_text = comment_element.find_element(By.XPATH,".//div[@dir='auto']").text

    sheet.cell(row=row, column=1, value=user_name)
    sheet.cell(row=row, column=2, value=comment_text)

    row += 1

# Guardar el archivo de Excel
workbook.save("comentarios.xlsx")

# Cierra el navegador al finalizar
driver.quit()

  
